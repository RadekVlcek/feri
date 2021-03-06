# TODO:
# - (DONE) consider bordering the area using thick border
# - (DONE) gather font family and font size as well
# - (DONE) fix width obtain when changing width to more columns at once
# - (DONE) add get font color function
# - gather info about vertical & horizontal align
# - find a way to get data from Theme colors palette

class Feri:
    def __init__(self, wb_path):
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.worksheet.dimensions import ColumnDimension

        # Worksheet initialization for Read-Only
        try:
            wb = load_workbook(filename=wb_path)
        except FileNotFoundError as err:
            raise err
        else:
            self.sheet = wb.active

        # Data storage
        self.excelData = {
            "boardData": {},
            "rowData": [],
            "columnData": []
        }

        # Number of rows and columns IN USE
        self.row_count = self.sheet.max_row
        self.col_count = self.sheet.max_column

        # GLOBAL variables
        # Get column width in pixels
        self.prev_width = 64

    def get_column_width(self, column):
        
        # 0.08 from 8
        # 0.17 from 9
        # 0.25 from 8
        # 0.33 from 8
        # 0.42 from 9
        # 0.50 from 8
        # 0.58 from 8
        # 0.67 from 9
        # 0.75 from 8
        # 0.83 from 8
        # 0.92 from 9
        # 1.00 from 8

        # 1.14 from 14
        # 1.29 from 15
        # 1.43 from 14
        # 1.57 from 14
        # 1.71 from 14
        # 1.86 from 15
        # 2.00 from 14
        
        # Difference between Excel points and python output = 0.71578125, sometimes 0.710625 and so on.

        #1      0.14453125
        #2      0.140625
        #3      0.14453125
        #4      0.140625
        #3      0.14453125
        #6      0.14453125
        #7      0.140625
        #8      0.14453125
        #9      0.140625
        #10     0.14453125

        #11     0.140625
        # reached 1.00 points here

        # I only need these two values: 
        # 0.140625 - starting value
        # 0.14453125 - second value

        # boha kym som toto vylustil...

        column_letters = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z")
        letter = column_letters[column]
        total_xl_points = self.sheet.column_dimensions[letter].width

        if total_xl_points == 0.0:
            return self.prev_width

        else:
            xl_points, w_vals = 0.0, (0.140625, 0.14453125)
            v = pixels = 0

            while xl_points < total_xl_points:
                xl_points = xl_points + w_vals[v]
                v = 1 - v
                pixels += 1

            width = pixels - 1
            self.prev_width = width

            return width

    def get_value(self, cell):
        return cell.value if cell.value is not None else ''

    def get_row_height(self, row):
        return int(row.height / 0.75) if row.height is not None else 20

    def get_final_rgb(self, hex_val):
        stripped = hex_val[2:len(hex_val)]
        return '#FFFFFF' if stripped == '000000' else f'#{stripped}'

    def get_color(self, cell):
        color = cell.font.color
        return self.get_final_rgb(color.rgb) if color.type == 'rgb' else '#000000'

    def get_bg_color(self, cell):
        color = cell.fill.start_color
        if color.type == 'rgb':
            return self.get_final_rgb(color.rgb)
        # THEME COLORS ISSUE. To be continued...

    def get_font_family(self, cell):
        return cell.font.name

    def get_font_size(self, cell):
        return cell.font.sz

    def get_font_weight(self, cell):
        return 'bold' if cell.font.b else 'normal'

    def get_font_style(self, cell):
        return 'italic' if cell.font.i else 'normal'

    def get_text_decoration_style(self, cell):
        return cell.font.u

    def get_text_decoration(self, cell):
        return 'line-through' if cell.font.strike else 'none'

    def get_text_align(self, cell):
        if cell.alignment.horizontal is None:
            return 'right' if type(cell.value) is int else 'left'
        else:
            return cell.alignment.horizontal

    def get_vertical_align(self, cell):
        if cell.alignment.vertical is None:
            return 'bottom'
        else:
            return 'top' if cell.alignment.vertical == 'top' else 'middle'

    def get_border(self, sheet, row, column):
        cell = sheet.cell(row=row, column=column)
        left_border = cell.border.left.style
        right_border = cell.border.right.style

        # print(f'left_border: {left_border}, right_border: {right_border}')

        if left_border is not None:
            print(f'{row}{column} - row merge start')
        
        elif right_border is not None:
            print(f'{row}{column} - row merge end')

    def saveData(self):

        total_board_height = 0

        for row in range(1, self.row_count+1):
            self.excelData['columnData'].append([])
            self.excelData['rowData'].append([])

            total_board_height += self.get_row_height(self.sheet.row_dimensions[row])

            total_row_width = 0
            total_row_height = 0

            for column in range(1, self.col_count+1):

                # exploring 
                # print(self.sheet.cell(row=row, column=column).alignment.vertical)

                total_row_width += self.get_column_width(column-1)
                total_row_height = self.get_row_height(self.sheet.row_dimensions[row])

                # Save parameters for each column
                self.excelData['columnData'][row-1].append({
                    "column": column,
                    "value": self.get_value(self.sheet.cell(row=row, column=column)),
                    "width": self.get_column_width(column-1),
                    "height": self.get_row_height(self.sheet.row_dimensions[row]),
                    "color": self.get_color(self.sheet.cell(row=row, column=column)),
                    "background_color": self.get_bg_color(self.sheet.cell(row=row, column=column)),
                    "font_family": self.get_font_family(self.sheet.cell(row=row, column=column)),
                    "font_size": self.get_font_size(self.sheet.cell(row=row, column=column)),
                    "font_weight": self.get_font_weight(self.sheet.cell(row=row, column=column)),
                    "font_style": self.get_font_style(self.sheet.cell(row=row, column=column)),
                    "text_decoration": self.get_text_decoration(self.sheet.cell(row=row, column=column)),
                    "text_decoration-style": self.get_text_decoration_style(self.sheet.cell(row=row, column=column)),
                    "text_align": self.get_text_align(self.sheet.cell(row=row, column=column)),
                    "vertical_align": self.get_vertical_align(self.sheet.cell(row=row, column=column)),
                    "border": self.get_border(self.sheet, row, column)
                })
            
            # Save parameters for each row
            self.excelData['rowData'][row-1].append({
                    "width": total_row_width,
                    "height": total_row_height
                })
        
        # Save parameters for the board
        self.excelData['boardData']['board_width'] = total_row_width
        self.excelData['boardData']['board_height'] = total_board_height