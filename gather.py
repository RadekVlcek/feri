# TODO:
# - consider bordering the area using thick border
# - fix Theme colors functionality 
# - (DONE) fix width obtain when changing width to more columns at once
# - (DONE) add get font color function

class Gather:

    def __init__(self, wb_path):
        import json
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.worksheet.dimensions import ColumnDimension

        # Worksheet initialization for Read-Only
        wb = load_workbook(filename = wb_path)
        self.sheet = wb.active

        # Data storage
        self.excelData = {
            "metadata": {},
            "data": []
        }

        # Number of rows and columns IN USE
        self.row_count = self.sheet.max_row
        self.col_count = self.sheet.max_column

        # GLOBAL variables
        # Get column width in pixels
        self.prev_width = 0

    def get_column_width(self, column):
        
        # 0.08 from 8
        # 0.17 from 9
        # 0.25 from 8
        # 0.33 from 8
        # 0.42 from 9
        # 0.50 from 8
        # 0.58 from 8
        # 0.67 from 9
        # 0.75 from 8
        # 0.83 from 8
        # 0.92 from 9
        # 1.00 from 8

        # 1.14 from 14
        # 1.29 from 15
        # 1.43 from 14
        # 1.57 from 14
        # 1.71 from 14
        # 1.86 from 15
        # 2.00 from 14
        
        # Difference between Excel points and python output = 0.71578125, sometimes 0.710625 and so on.

        #1      0.14453125
        #2      0.140625
        #3      0.14453125
        #4      0.140625
        #3      0.14453125
        #6      0.14453125
        #7      0.140625
        #8      0.14453125
        #9      0.140625
        #10     0.14453125

        #11     0.140625
        # reached 1.00 points here

        # I only need these two values: 
        # 0.140625 - starting value
        # 0.14453125 - second value

        # boha kym som toto vylustil...

        column_letters = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z")
        letter = column_letters[column]
        total_xl_points = self.sheet.column_dimensions[letter].width

        if total_xl_points == 0.0:
            return self.prev_width

        else:
            # Pripocitavaju sa tieto dve hodnoty - striedavo
            w_vals = (0.140625, 0.14453125)
            xl_points = 0.0
            pixels = 0
            v = 0
            while xl_points < total_xl_points:
                xl_points = xl_points + w_vals[v]
                v = 1 - v
                pixels += 1

            width = pixels - 1
            self.prev_width = width

            return width

    def get_value(self, value):
        if value is not None:
            return value
        else:
            return 'empty'

    def get_row_height(self, row):
        height = self.sheet.row_dimensions[row].height
        if height is not None:
            return int(height / 0.75)
        else:
            return 20

    def get_final_rgb(self, hex_val):
        stripped = hex_val[2:len(hex_val)]
        if stripped == '000000':
            return '#FFFFFF'
        else:
            return f'#{stripped}'

    def get_color(self, cell):
        color = cell.font.color
        if color.type == 'rgb':
            return self.get_final_rgb(color.rgb)
        else:
            return '#000000'

    def get_bg_color(self, cell):
        color = cell.fill.start_color
        if color.type == 'rgb':
            return self.get_final_rgb(color.rgb)

    def get_border(self, sheet, row, column):
        cell = sheet.cell(row=row, column=column)
        left_border = cell.border.left.style
        right_border = cell.border.right.style

        # print(f'left_border: {left_border}, right_border: {right_border}')

        if left_border is not None:
            print(f'{row}{column} - row merge start')
        
        elif right_border is not None:
            print(f'{row}{column} - row merge end')

    def saveData(self):
        self.excelData['metadata']['row_count'] = self.row_count
        self.excelData['metadata']['col_count'] = self.col_count

        for row in range(1, self.row_count+1):
            for column in range(1, self.col_count+1):
                self.excelData['data'].append({
                    "row": row,
                    "column": column,
                    "value": self.get_value(self.sheet.cell(row=row, column=column).value),
                    "height": self.get_row_height(row),
                    "width": self.get_column_width(column-1),
                    "color": self.get_color(self.sheet.cell(row=row, column=column)),
                    "bg_color": self.get_bg_color(self.sheet.cell(row=row, column=column)),
                    "border": self.get_border(self.sheet, row, column)
                })